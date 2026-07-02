from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def default_source_dir() -> Path:
    return Path(__file__).resolve().parents[2] / "测试场站数据" / "nmg_shuijinghu_solar"


def load_station_info(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = workbook.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        workbook.close()
    station_name = None
    lon = None
    lat = None
    capacity_mw = None
    for row in rows:
        if not row:
            continue
        if row[1]:
            station_name = str(row[1]).strip()
        key = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        value = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        if "经纬度" in key:
            nums = [float(x) for x in re.findall(r"-?\d+(?:\.\d+)?", value)]
            if len(nums) >= 2:
                lon, lat = nums[0], nums[1]
        if "装机容量" in key and capacity_mw is None:
            match = re.search(r"-?\d+(?:\.\d+)?", value)
            if match:
                capacity_mw = float(match.group(0))
    return {
        "station_id": "nmg_shuijinghu_solar",
        "station_name": station_name or "水镜湖光伏电站",
        "longitude": lon,
        "latitude": lat,
        "capacity_mw": capacity_mw,
        "source_file": str(path),
    }


def prepare(source_dir: Path) -> dict:
    power_path = source_dir / "station_power" / "station_power_original.csv"
    station_path = source_dir / "station_info" / "水镜湖光伏电站收资表.xlsx"
    output_path = source_dir / "station_power" / "nmg_shuijinghu_solar_real.csv"
    station_info_path = source_dir / "station_info" / "nmg_shuijinghu_station_info.txt"
    summary_path = source_dir / "station_power" / "nmg_shuijinghu_prepare_summary.json"

    raw = pd.read_csv(power_path)
    frame = pd.DataFrame(
        {
            "bj_time": pd.to_datetime(raw["bj_time"], errors="coerce"),
            "actual_power": pd.to_numeric(raw["actual_power"], errors="coerce"),
            "direct_irradiance": pd.to_numeric(raw["radiation_total"], errors="coerce"),
        }
    )
    frame = frame.dropna(subset=["bj_time"]).sort_values("bj_time").drop_duplicates("bj_time", keep="last")
    frame["bj_time"] = frame["bj_time"].dt.strftime("%Y-%m-%d %H:%M:%S")
    frame.to_csv(output_path, index=False, encoding="utf-8-sig")

    station = load_station_info(station_path)
    station_info_path.write_text(json.dumps(station, indent=2, ensure_ascii=False), encoding="utf-8")
    summary = {
        "source_power": str(power_path),
        "output_csv": str(output_path),
        "station_info": str(station_info_path),
        "rows_raw": int(len(raw)),
        "rows_output": int(len(frame)),
        "start_time": str(frame["bj_time"].min()) if len(frame) else None,
        "end_time": str(frame["bj_time"].max()) if len(frame) else None,
        "actual_power_missing_rate": float(pd.to_numeric(frame["actual_power"], errors="coerce").isna().mean()) if len(frame) else 1.0,
        "direct_irradiance_missing_rate": float(pd.to_numeric(frame["direct_irradiance"], errors="coerce").isna().mean()) if len(frame) else 1.0,
        "station": station,
    }
    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    return summary


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", default=str(default_source_dir()))
    args = parser.parse_args()
    summary = prepare(Path(args.source_dir))
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
